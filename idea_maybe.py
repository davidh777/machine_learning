import tensorflow as tf
from tensorflow import keras
import numpy as np
import random
import openpyxl as xl
import time as t

data = keras.datasets.imdb
wb = xl.load_workbook("idea.xlsx")
sheet = wb["Sheet1"]
(train_data, train_labels), (test_data, test_labels) = data.load_data(num_words=10000)

word_index = data.get_word_index()
word_index = {k: (v+3) for k, v in word_index.items()}
word_index["<PAD>"] = 0
word_index["<START>"] = 1
word_index["<UNK>"] = 2
word_index["<UNUSED>"] = 3

reverse_word_index = dict([(value, key) for (key, value) in word_index.items()])
train_data = keras.preprocessing.sequence.pad_sequences(train_data, value=word_index["<PAD>"], padding="post", maxlen=250)
test_data = keras.preprocessing.sequence.pad_sequences(test_data, value=word_index["<PAD>"], padding="post", maxlen=250)

activation_function = {
    1: "relu",
    2: "sigmoid",
    3: "tanh"
}


def decode_review(text):
    return " ".join([reverse_word_index.get(i, "?") for i in text])


for z in range(2, 300):
    start = t.time()
    n = random.randint(2, 512)
    e = random.randint(2, 82)
    b = random.randint(10, 600)
    na = random.randint(1, 100)
    nb = random.randint(1, 100)
    a = random.randint(1, 3)
    a2 = random.randint(1, 3)
    a3 = random.randint(1, 3)
    a4 = random.randint(1, 3)
    model = keras.Sequential()
    model.add(keras.layers.Embedding(10000, 16))
    model.add(keras.layers.GlobalAveragePooling1D())
    model.add(keras.layers.Dense(n, activation=activation_function[a]))
    if na != 0:
        model.add(keras.layers.Dense(na, activation=activation_function[a2]))
    if nb != 0:
        model.add(keras.layers.Dense(nb, activation=activation_function[a3]))
    model.add(keras.layers.Dense(1, activation=activation_function[a4]))
    model.summary()
    model.compile(optimizer="adam", loss="binary_crossentropy", metrics=["accuracy"])
    x_val = train_data[:10000]
    x_train = train_data[10000:]
    y_val = train_labels[:10000]
    y_train = train_labels[10000:]

    fitModel = model.fit(x_train, y_train, epochs=e, batch_size=b, validation_data=(x_val, y_val), verbose=1)
    results = model.evaluate(test_data, test_labels)
    end = t.time()
    cell_cost = sheet.cell(z, 4)
    cell_acc = sheet.cell(z, 3)
    cell_cost.value = results[0]
    cell_acc.value = results[1]
    cell_epochs = sheet.cell(z, 1)
    cell_epochs.value = e
    cell_neurons = sheet.cell(z, 2)
    cell_neurons.value = n
    cell_batch = sheet.cell(z, 5)
    cell_batch.value = b
    cell_time = sheet.cell(z, 6)
    cell_time.value = end - start
    cell_na = sheet.cell(z, 7)
    cell_na.value = na
    cell_act1 = sheet.cell(z, 8)
    cell_act1.value = a
    cell_act2 = sheet.cell(z, 9)
    cell_act2.value = a2
    cell_act3 = sheet.cell(z, 10)
    cell_act3.value = a3
    cell_act4 = sheet.cell(z, 11)
    cell_act4.value = a4
    cell_nb = sheet.cell(z, 12)
    cell_nb.value = nb
wb.save("data_seta.xlsx")
