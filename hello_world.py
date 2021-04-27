import tensorflow as tf
from tensorflow import keras
import numpy as np
import matplotlib.pyplot as plt
import time
import openpyxl as xl
import random as random

wb = xl.load_workbook("idea.xlsx")
sheet = wb["Sheet1"]
start = time.time()
data = keras.datasets.fashion_mnist

(train_images, train_labels), (test_images, test_labels) = data.load_data()
best = 0

class_names = ['T-Shirt/top', 'Trouser', 'Pullover', 'Dress', 'Coat', 'Sandal', 'Shirt', 'Sneaker', 'Bag', 'Ankle boot'] # Different classifcations
(train_images, train_labels), (test_images, test_labels) = data.load_data()

train_images = train_images/255.0
test_images = test_images/255.0
for n in range(2, 3):
    model = keras.Sequential([
        keras.layers.Flatten(input_shape=(28, 28)),
        keras.layers.Dense(128, activation="relu"),
        keras.layers.Dense(10, activation="softmax")
    ])
    z = random.randint(2, 28)
    model.compile(optimizer="adam", loss="sparse_categorical_crossentropy", metrics=["accuracy"])
    model.fit(train_images, train_labels, epochs=z)

    test_loss, test_acc = model.evaluate(test_images, test_labels)
    end = time.time()
    cell_epochs = sheet.cell(n, 1)
    cell_epochs.value = z
    cell_time = sheet.cell(n, 2)
    cell_time.value = end - start
    cell_acc = sheet.cell(n, 3)
    cell_acc.value = test_acc

